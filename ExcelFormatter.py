import openpyxl as oxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from itertools import islice
import re


class ExcelFormatter(object):
    def __init__(self, filename):
        self.filename = filename
        self.wb = oxl.load_workbook(filename)
        self.ws = self.wb.active
        self.ProcessColumnIndices()

        self.color_red = "FFFF0000"
        self.color_yellow = "FFFFFF00"
        self.color_white = "FFFFFFFF"

    def SaveWorkbook(self, new_filename):
        self.wb.save(new_filename)

    def ProcessColumnIndices(self):
        self.indices = {}
        n = 0
        for _, first_row in enumerate(islice(self.ws.iter_rows(), 1)):
            for cell in first_row:
                self.indices[cell.value] = n
                n += 1

    def GetColumnIndicesByName(self, column_names):
        indices = []
        for c_name in column_names:
            indices.append(self.indices.get(c_name))
        return indices

    def CheckAddressLines(self, column_names):
        errors = 0
        for idx in self.GetColumnIndicesByName(column_names):
            for _, selected_column in enumerate(islice(self.ws.iter_cols(), idx, idx+1)):
                for cell in islice(selected_column, 2, None):
                    if cell.value is None:
                        self.fillCellColour(cell, self.color_red)
                        errors += 1
                    elif not cell.value.strip().endswith(","):
                        cell.value = cell.value.strip() + ","
                    else:
                        cell.value = cell.value.strip()
        return errors

    def CheckEmailAddressLines(self, column_names):
        errors = 0
        for idx in self.GetColumnIndicesByName(column_names):
            for _, selected_column in enumerate(islice(self.ws.iter_cols(), idx, idx+1)):
                for cell in islice(selected_column, 2, None):
                    if cell.value is None:
                        self.fillCellColour(cell, self.color_red)
                        errors += 1
                    elif self.isNotAvailable(cell.value.strip()):
                        cell.value = "NO"
                    elif not re.match(r"[^@]+@[^@]+\.[^@]+", cell.value.strip()) and cell.value.strip() != "NO":
                        self.fillCellColour(cell, self.color_yellow)
                        errors += 1
                    else:
                        cell.value = cell.value.strip()
        return errors

    def CheckPhoneNumbersLines(self, column_names):
        errors = 0
        for idx in self.GetColumnIndicesByName(column_names):
            for _, selected_column in enumerate(islice(self.ws.iter_cols(), idx, idx+1)):
                for cell in islice(selected_column, 2, None):
                    if cell.value is None:
                        self.fillCellColour(cell, self.color_red)
                        errors += 1
                    elif self.isNotAvailable(cell.value.strip()):
                        cell.value = "NO"
                    elif not re.match(r"^(\+\d{1,3}\ \d{10})$", cell.value.strip()) and cell.value.strip() != "NO":
                        # Replacing invisible non-really-a-space
                        new_str = str(cell.value).replace(" ", "-")
                        p = re.compile(r"^(\+\d{1,3}\-\d{10})$")
                        found_val = p.search(new_str)
                        if found_val is not None:
                            cell.value = found_val.group(0).replace("-", " ")
                        else:
                            self.fillCellColour(cell, self.color_yellow)
                            errors += 1
                    else:
                        cell.value = cell.value.strip()
        return errors

    def CheckTitleLines(self, column_names):
        errors = 0
        for idx in self.GetColumnIndicesByName(column_names):
            for _, selected_column in enumerate(islice(self.ws.iter_cols(), idx, idx+1)):
                for cell in islice(selected_column, 2, None):
                    if cell.value is None:
                        self.fillCellColour(cell, self.color_red)
                        errors += 1
                    elif not (cell.value.strip() in ["Mr.", "Dr.", "Mrs.", "Ms.", "Adv."]):
                        if "mrs" in cell.value.strip().lower():
                            cell.value = "Mrs."
                        elif "mr" in cell.value.strip().lower():
                            cell.value = "Mr."
                        elif "ms" in cell.value.strip().lower():
                            cell.value = "Ms."
                        elif "dr" in cell.value.strip().lower():
                            cell.value = "Dr."
                        elif "adv" in cell.value.strip().lower():
                            cell.value = "Adv."
                        else:
                            self.fillCellColour(cell, self.color_yellow)
                            errors += 1
                    else:
                        cell.value = cell.value.strip()
        return errors

    def CheckLatLonLines(self, column_names):
        errors = 0
        for idx in self.GetColumnIndicesByName(column_names):
            for _, selected_column in enumerate(islice(self.ws.iter_cols(), idx, idx+1)):
                for cell in islice(selected_column, 2, None):
                    if cell.value is None:
                        self.fillCellColour(cell, self.color_red)
                        errors += 1
                    elif not isinstance(cell.value, float) and self.isNotAvailable(cell.value.strip()):
                        cell.value = "NI"
                    elif not re.match(r"^(\d{1,2}.\d{4,10} \d{1,3}.\d{4,10})$", str(cell.value)) and str(cell.value) != "NI":
                        self.fillCellColour(cell, self.color_yellow)
                        errors += 1
        return errors

    def CheckHoursLines(self, column_names):
        errors = 0
        for idx in self.GetColumnIndicesByName(column_names):
            for _, selected_column in enumerate(islice(self.ws.iter_cols(), idx, idx+1)):
                for cell in islice(selected_column, 2, None):
                    if cell.value is None:
                        self.fillCellColour(cell, self.color_red)
                        errors += 1
                    elif not re.match(r"^([01]?[0-9]|2[0-3]):[0-5][0-9]$", str(cell.value)):
                        p = re.compile(r"([01]?[0-9]|2[0-3]):[0-5][0-9]")
                        found_val = p.search(str(cell.value))
                        if found_val is not None:
                            cell.value = found_val.group(0)
                        else:
                            self.fillCellColour(cell, self.color_yellow)
                            errors += 1
        return errors

    def CheckNumberRangeLines(self, column_names):
        errors = 0
        indices = self.GetColumnIndicesByName(column_names)
        column_number_idx = indices[0]
        column_range_idx = indices[1]
        cell_index = 3
        for _, selected_column in enumerate(islice(self.ws.iter_cols(), column_number_idx, column_number_idx+1)):
            for cell in islice(selected_column, 2, None):
                range_cell = self.ws.cell(
                    row=cell_index, column=column_range_idx + 1)
                errors += self.CheckCountOrRange(cell, range_cell)
                cell_index += 1
        return errors

    def CheckCountOrRange(self, cell_number, cell_range):
        if cell_number.value is None:
            if cell_range.value is None:
                self.fillCellColour(cell_number, self.color_red)
                self.fillCellColour(cell_range, self.color_red)
                return 1
            else:
                if not re.match(r"^[']?\d{1,10}-\d{1,10}$", str(cell_range.value).strip()) and str(cell_range.value).strip() != "NA":
                    self.fillCellColour(cell_range, self.color_yellow)
                    return 1
        elif not re.match(r"^\d{1,10}$", str(cell_number.value).strip()):
            self.fillCellColour(cell_number, self.color_yellow)
            return 1
        return 0

    def CheckSpecialitiesLines(self, column_names):
        errors = 0
        indices = self.GetColumnIndicesByName(column_names)
        column_spec_idx = indices[0]
        column_number_idx = indices[1]
        column_range_idx = indices[2]
        cell_index = 3
        for _, selected_column in enumerate(islice(self.ws.iter_cols(), column_spec_idx, column_spec_idx+1)):
            for cell in islice(selected_column, 2, None):
                number_cell = self.ws.cell(
                    row=cell_index, column=column_number_idx + 1)
                range_cell = self.ws.cell(
                    row=cell_index, column=column_range_idx + 1)
                if str(cell.value).strip().upper() == "YY":
                    cell.value = "YY"
                    errors += self.CheckCountOrRange(number_cell, range_cell)
                elif str(cell.value).strip().upper() == "NA" or str(cell.value).strip().upper() == "NO":
                    cell.value = "NA"
                    if number_cell.value is not None:
                        number_cell.value = ""
                    if str(range_cell.value).strip() != "NA":
                        range_cell.value = "NA"
                else:
                    self.fillCellColour(cell, self.color_red)
                cell_index += 1
        return errors

    def CheckTotalLines(self, total_column_names, column_names, range_totals):
        errors = 0
        total_column_number_idx = self.GetColumnIndicesByName(total_column_names)[
            0]
        total_column_range_idx = self.GetColumnIndicesByName(total_column_names)[
            1]
        cell_index = 3
        for _, selected_column in enumerate(islice(self.ws.iter_cols(), total_column_number_idx, total_column_number_idx+1)):
            for cell in islice(selected_column, 2, None):
                total = 0
                total_max_range = 0
                is_range = False

                for column_pair in column_names:
                    column_number_idx = self.GetColumnIndicesByName(column_pair)[
                        0]
                    column_range_idx = self.GetColumnIndicesByName(column_pair)[
                        1]

                    number_cell = self.ws.cell(
                        row=cell_index, column=column_number_idx + 1)
                    range_cell = self.ws.cell(
                        row=cell_index, column=column_range_idx + 1)

                    if number_cell.value is None:
                        if re.match(r"^[']?\d{1,10}-\d{1,10}$", str(range_cell.value).strip()) and str(range_cell.value).strip() != "NA":
                            range_value = str(range_cell.value).strip().replace(
                                "'", "").split("-")
                            total_max_range += int(range_value[1])
                            is_range = True
                    elif re.match(r"^\d{1,10}$", str(number_cell.value).strip()):
                        total += int(number_cell.value)

                range_cell = self.ws.cell(
                    row=cell_index, column=total_column_range_idx + 1)
                if is_range:
                    selected_range = self.GetCorrectTotalRange(
                        total_max_range, range_totals)
                    cell.value = ""
                    range_cell.value = selected_range
                    self.fillCellColour(cell, self.color_white)
                    self.fillCellColour(range_cell, self.color_white)
                else:
                    cell.value = total
                    range_cell.value = ""
                    self.fillCellColour(cell, self.color_white)
                    self.fillCellColour(range_cell, self.color_white)
                cell_index += 1
        return errors

    def GetCorrectTotalRange(self, total_max_range, range_totals):
        for r in range_totals:
            # Get Range Max
            if r.endswith("<"):
                val = int(r.split("<")[0])
            else:
                val = int(r.split("-")[1])

            if total_max_range <= val:
                return r
        return range_totals[len(range_totals)-1]

    def CheckYYNOAPPLines(self, column_names):
        errors = 0
        for idx in self.GetColumnIndicesByName(column_names):
            for _, selected_column in enumerate(islice(self.ws.iter_cols(), idx, idx+1)):
                for cell in islice(selected_column, 2, None):
                    if cell.value is None:
                        self.fillCellColour(cell, self.color_red)
                        errors += 1
                    elif cell.value.strip().lower() == "yy" or cell.value.strip().lower() == "app":
                        cell.value = "YY"
                    elif cell.value.strip().lower() == "no":
                        cell.value = "NO"
                    else:
                        self.fillCellColour(cell, self.color_yellow)
                        errors += 1
        return errors

    def CheckAllowedName(self, column_names, state_names):
        errors = 0
        for idx in self.GetColumnIndicesByName(column_names):
            for _, selected_column in enumerate(islice(self.ws.iter_cols(), idx, idx+1)):
                for cell in islice(selected_column, 2, None):
                    if cell.value is None:
                        self.fillCellColour(cell, self.color_red)
                        errors += 1
                    elif cell.value.strip() in state_names:
                        cell.value = cell.value.strip()
                    else:
                        self.fillCellColour(cell, self.color_yellow)
                        errors += 1
        return errors

    def fillCellColour(self, cell, colour):
        cell.fill = PatternFill(
            start_color=colour, end_color=colour, fill_type="solid")

    def isNotAvailable(self, value):
        return value == "NO" or value == "NI" or value == "NA" or value == "N0" or value == "NP"

    def CheckDuplicateHospitals(self, list_hospital_info):
        errors = 0
        hospital_names = []
        hospital_addr1 = []
        hospital_addr2 = []
        hospital_addr3 = []
        hospital_city = []
        hospital_state = []
        indices = self.GetColumnIndicesByName(list_hospital_info)
        hospital_name_idx = indices[0]
        hospital_addr1_idx = indices[1]
        hospital_addr2_idx = indices[2]
        hospital_addr3_idx = indices[3]
        hospital_city_idx = indices[4]
        hospital_state_idx = indices[5]
        cell_index = 3
        for _, selected_column in enumerate(islice(self.ws.iter_cols(), hospital_name_idx, hospital_name_idx+1)):
            for cell in islice(selected_column, 2, None):
                addr1_cell = self.ws.cell(
                    row=cell_index, column=hospital_addr1_idx + 1)
                addr2_cell = self.ws.cell(
                    row=cell_index, column=hospital_addr2_idx + 1)
                addr3_cell = self.ws.cell(
                    row=cell_index, column=hospital_addr3_idx + 1)
                city_cell = self.ws.cell(
                    row=cell_index, column=hospital_city_idx + 1)
                state_cell = self.ws.cell(
                    row=cell_index, column=hospital_state_idx + 1)
                duplicate = False
                try:
                    index = hospital_names.index(cell.value)
                except ValueError:
                    index = -1

                # Hospital Name Found
                if index is not -1:
                    duplicate = (addr1_cell == hospital_addr1[index] and
                                 addr2_cell == hospital_addr2[index] and
                                 addr3_cell == hospital_addr3[index] and
                                 city_cell == hospital_city[index] and
                                 state_cell == hospital_state[index])

                if not duplicate:
                    hospital_names.append(cell.value)
                    hospital_addr1.append(addr1_cell.value)
                    hospital_addr2.append(addr2_cell.value)
                    hospital_addr3.append(addr3_cell.value)
                    hospital_city.append(city_cell.value)
                    hospital_state.append(state_cell.value)
                else:
                    self.fillCellColour(cell, self.color_red)
                    self.fillCellColour(addr1_cell, self.color_red)
                    self.fillCellColour(addr2_cell, self.color_red)
                    self.fillCellColour(addr3_cell, self.color_red)
                    self.fillCellColour(city_cell, self.color_red)
                    self.fillCellColour(state_cell, self.color_red)
                    errors += 1
                cell_index += 1
        return errors
